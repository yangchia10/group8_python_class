import requests
from bs4 import BeautifulSoup
import time
import openpyxl

wb = openpyxl.Workbook()    #workbook
ws = wb.active              #worksheet

page = 1

while True:
    res = requests.get('https://www.104.com.tw/jobs/search/?ro=0&kwop=7&keyword=%E5%A4%A7%E6%95%B8%E6%93%9A&expansionType=area%2Cspec%2Ccom%2Cjob%2Cwf%2Cwktm&order=15&asc=0&page=' + str(page) + '&mode=s&jobsource=2018indexpoc&langFlag=0&langStatus=0&recommendJob=1&hotJob=1')
    soup = BeautifulSoup(res.text)

    jobs = soup.find_all('article', class_='b-block--top-bord job-list-item b-clearfix js-job-item')

    if not jobs:
        break

    print("----------------------")
    print("現在正在讀取第" + str(page) + "頁")
    print("----------------------")

    for job in jobs:
        
        a = job['data-job-name'].strip()
        b = 'https:' + job.find("a", class_='js-job-link')['href']
        c = job('li')[2].text.strip()
        d = job('ul')[1].text.strip()        #c, d 有錯

        job_tag = job.find('div', class_="job-list-tag b-content")
        if job_tag.select('span') and job_tag.select('span')[0].text == "待遇面議":   #加上&後面的條件是因為，遠端工作也是tag span
            e = job_tag.span.text
            f = ""
            g = ""
            h = ""
        else:
            e = job_tag.a.text
            ###
            #excel中的f欄 = e欄取前2個字
            f = e[:2]                               #取前兩個字，!!!!!起點包含終點不包含!!!!!

            #刪除無用字元
            salary = ""
            for char in e:
                if char.isdigit() or char == "~":
                    salary += char

            if "~" in e:    
                g = salary[:salary.find('~')]       #找下限    
                h = salary[salary.find('~')+1:]     #找上限
            else:
                g = salary
                h = salary
            ###
            ###將產出f,g,h作為output

        try:
            ws.append([a, b, c, d, e, f, g, h])
        except:
            print(a)
            print(b)
            print(c)
            print(d)
            print(e)
            print(f)
            print(g)
            print(h)
            print("error here")
            break
    

    page += 1

    time.sleep(1)

    ws['A1'] = "職缺名稱"
    ws['B1'] = "職缺連結"
    ws['C1'] = "公司名稱"
    ws['D1'] = "工作地區"
    ws['E1'] = "薪資待遇"
    ws['F1'] = "給薪方式"
    ws['G1'] = "薪資下限"
    ws['H1'] = "薪資上限"

    wb.save("excel測試20230602.xlsx")