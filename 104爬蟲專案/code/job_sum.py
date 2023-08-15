import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 讀取Excel文件
wb = load_workbook('group8_python_class\\104爬蟲專案\\計算資料數量.xlsx')
sheet = wb['sss']

# 將數據讀取到DataFrame中
data = sheet.values
df = pd.DataFrame(data)

# 定義要處理的列索引
columns = [34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48]

# 過濾掉空值並清理標準化數據
for col in columns:
    df[col] = df[col].str.strip().str.lower()

# 計算數據的出現次數
combined_value_counts = pd.Series(dtype=int)  # 初始化結果為空的Series
for col in columns:
    value_counts = df[col].value_counts()
    combined_value_counts = combined_value_counts.add(value_counts, fill_value=0)

# 輸出結果
print(combined_value_counts.astype(int))

# 將結果寫入新的工作表
summary_sheet = wb.create_sheet('工作技能')

# 將結果轉換為DataFrame並寫入工作表
summary_df = pd.DataFrame({'Value': combined_value_counts.index, 'Count': combined_value_counts.astype(int)})
for row in dataframe_to_rows(summary_df, index=False, header=True):
    summary_sheet.append(row)

# 保存更改
wb.save('group8_python_class\\104爬蟲專案\\計算資料數量.xlsx')






